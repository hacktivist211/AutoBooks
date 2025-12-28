from pathlib import Path
from typing import Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.models import Transaction
from src.logger import get_logger

logger = get_logger(__name__)

class ExcelLedger:
    """Excel ledger writer for Tally-compatible transaction output."""

    # Column definitions
    COLUMNS = [
        "Date",
        "Vendor",
        "Debit Account",
        "Debit Amount",
        "Credit Account",
        "Credit Amount",
        "TDS Account",
        "TDS Amount",
        "Confidence",
        "Status"
    ]

    def __init__(self, filepath: str):
        """
        Initialize Excel ledger writer.

        Args:
            filepath: Path to Excel file
        """
        self.filepath = Path(filepath)
        self.filepath.parent.mkdir(parents=True, exist_ok=True)

        # Load or create workbook
        if self.filepath.exists():
            logger.info(f"Loading existing Excel file: {self.filepath}")
            self.workbook = openpyxl.load_workbook(self.filepath)
        else:
            logger.info(f"Creating new Excel file: {self.filepath}")
            self.workbook = openpyxl.Workbook()

        # Get or create sheet
        if "Transactions" in self.workbook.sheetnames:
            self.sheet = self.workbook["Transactions"]
        else:
            self.sheet = self.workbook.active
            self.sheet.title = "Transactions"
            self._create_headers()

        logger.info("ExcelLedger initialized")

    def _create_headers(self):
        """Create column headers with formatting."""
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Add headers
        for col_num, header in enumerate(self.COLUMNS, 1):
            cell = self.sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Set column widths
        column_widths = {
            "Date": 12,
            "Vendor": 25,
            "Debit Account": 20,
            "Debit Amount": 15,
            "Credit Account": 20,
            "Credit Amount": 15,
            "TDS Account": 15,
            "TDS Amount": 12,
            "Confidence": 12,
            "Status": 15
        }

        for col_num, header in enumerate(self.COLUMNS, 1):
            width = column_widths.get(header, 15)
            self.sheet.column_dimensions[get_column_letter(col_num)].width = width

    def append_transaction(self, txn: Transaction):
        """
        Append a transaction to the Excel sheet.

        Args:
            txn: Transaction object to add
        """
        # Get next row
        next_row = self.sheet.max_row + 1

        # Prepare row data
        row_data = [
            txn.date,
            txn.vendor,
            txn.debit_account,
            txn.debit_amount,
            txn.credit_account,
            txn.credit_amount,
            txn.tds_account,
            txn.tds_amount,
            f"{txn.confidence:.1%}",
            txn.status
        ]

        # Add data to row
        for col_num, value in enumerate(row_data, 1):
            cell = self.sheet.cell(row=next_row, column=col_num, value=value)

            # Format amounts as currency
            if col_num in [4, 6, 8]:  # Amount columns
                if value and isinstance(value, (int, float)):
                    cell.number_format = '₹#,##0.00'

            # Apply status-based coloring
            if txn.status == "AUTO_POSTED":
                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Light green
            elif txn.status == "USER_CONFIRMED":
                cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # Light yellow
            elif txn.status == "PATTERN_MATCHED":
                cell.fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")  # Light blue

        # Save workbook
        self.workbook.save(self.filepath)
        logger.info(f"Appended transaction for {txn.vendor} to Excel file")

    def get_summary(self) -> Dict[str, Any]:
        """
        Calculate summary statistics from the ledger.

        Returns:
            Dict with totals and statistics
        """
        total_debit = 0.0
        total_credit = 0.0
        total_tds = 0.0
        transaction_count = 0
        status_counts = {}

        # Skip header row
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Empty row
                continue

            transaction_count += 1

            # Sum amounts (handle None values)
            debit_amt = row[3] if row[3] else 0.0
            credit_amt = row[5] if row[5] else 0.0
            tds_amt = row[7] if row[7] else 0.0

            total_debit += float(debit_amt) if isinstance(debit_amt, (int, float)) else 0.0
            total_credit += float(credit_amt) if isinstance(credit_amt, (int, float)) else 0.0
            total_tds += float(tds_amt) if isinstance(tds_amt, (int, float)) else 0.0

            # Count statuses
            status = row[9] or "unknown"
            status_counts[status] = status_counts.get(status, 0) + 1

        summary = {
            "total_transactions": transaction_count,
            "total_debit": total_debit,
            "total_credit": total_credit,
            "total_tds": total_tds,
            "status_breakdown": status_counts
        }

        logger.info(f"Ledger summary: {transaction_count} transactions, ₹{total_debit:,.2f} debit, ₹{total_credit:,.2f} credit")
        return summary
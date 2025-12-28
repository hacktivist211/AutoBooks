import hashlib
from pathlib import Path
from typing import Optional
from PIL import Image
import pytesseract
import pdf2image
import openpyxl
from models import DocumentMetadata
from logger import get_logger
from datetime import datetime
import json

logger = get_logger(__name__)

class DocumentDecoder:
    """Handles decoding and text extraction from various document formats."""
    
    @staticmethod
    def get_file_hash(file_path: Path) -> str:
        """Generate hash of file for deduplication."""
        hasher = hashlib.sha256()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hasher.update(chunk)
        return hasher.hexdigest()
    
    @staticmethod
    def decode_pdf(file_path: Path) -> str:
        """Extract text from PDF using OCR."""
        try:
            logger.info(f"Decoding PDF: {file_path}")
            images = pdf2image.convert_from_path(file_path)
            extracted_text = ""
            
            for page_num, image in enumerate(images):
                logger.debug(f"Processing page {page_num + 1}")
                # Use Tesseract for OCR
                page_text = pytesseract.image_to_string(image)
                extracted_text += f"\n--- Page {page_num + 1} ---\n{page_text}"
            
            logger.info(f"Successfully extracted text from {len(images)} pages")
            return extracted_text
        except Exception as e:
            logger.error(f"Error decoding PDF {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def decode_excel(file_path: Path) -> str:
        """Extract text from Excel file."""
        try:
            logger.info(f"Decoding Excel: {file_path}")
            workbook = openpyxl.load_workbook(file_path)
            extracted_text = ""
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                extracted_text += f"\n--- Sheet: {sheet_name} ---\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    extracted_text += f"{row_text}\n"
            
            logger.info(f"Successfully extracted text from {len(workbook.sheetnames)} sheets")
            return extracted_text
        except Exception as e:
            logger.error(f"Error decoding Excel {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def decode_document(file_path: Path) -> tuple[str, DocumentMetadata]:
        """Decode document and return text + metadata."""
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        suffix = file_path.suffix.lower()
        
        if suffix == ".pdf":
            text = DocumentDecoder.decode_pdf(file_path)
            doc_type = "pdf"
        elif suffix == ".xlsx":
            text = DocumentDecoder.decode_excel(file_path)
            doc_type = "excel"
        elif suffix == ".txt":
            # For text files, just read as-is
            with open(file_path, "r", encoding="utf-8") as f:
                text = f.read()
            doc_type = "text"
        else:
            raise ValueError(f"Unsupported format: {suffix}")
        
        # Create metadata
        stat = file_path.stat()
        metadata = DocumentMetadata(
            document_id=f"{file_path.stem}_{int(stat.st_mtime)}",
            source_path=str(file_path),
            document_type=doc_type,
            created_at=datetime.fromtimestamp(stat.st_ctime),
            modified_at=datetime.fromtimestamp(stat.st_mtime),
            hash_value=DocumentDecoder.get_file_hash(file_path)
        )
        
        return text, metadata

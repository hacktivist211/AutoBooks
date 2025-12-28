from pathlib import Path
from typing import Optional
from datetime import datetime
from models import TransactionEntry
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from logger import get_logger

logger = get_logger(__name__)

class ExcelLedger:
    """Manages live Excel ledger output."""
    
    # Column definitions
    COLUMNS = [
        "Date",
        "Transaction ID",
        "Invoice ID",
        "Vendor",
        "Description",
        "Debit Account",
        "Debit Amount",
        "Credit Account",
        "Credit Amount",
        "TDS Account",
        "TDS Amount",
        "GST Amount",
        "Confidence",
        "Rule Applied",
        "Status"
    ]
    
    def __init__(self, output_path: Path):
        self.output_path = output_path
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self._initialize_workbook()
    
    def _initialize_workbook(self):
        """Create or open workbook."""
        if self.output_path.exists():
            logger.info(f"Opening existing workbook: {self.output_path}")
        else:
            logger.info(f"Creating new workbook: {self.output_path}")
            # Create new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Transactions"
            
            # Add headers with styling
            for col_idx, col_name in enumerate(self.COLUMNS, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set column widths
            column_widths = {
                "A": 12,  # Date
                "B": 20,  # Transaction ID
                "C": 15,  # Invoice ID
                "D": 20,  # Vendor
                "E": 25,  # Description
                "F": 18,  # Debit Account
                "G": 15,  # Debit Amount
                "H": 18,  # Credit Account
                "I": 15,  # Credit Amount
                "J": 15,  # TDS Account
                "K": 12,  # TDS Amount
                "L": 12,  # GST Amount
                "M": 12,  # Confidence
                "N": 20,  # Rule Applied
                "O": 12,  # Status
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            wb.save(self.output_path)
            logger.info(f"Workbook created with headers")
    
    def append_transaction(self, transaction: TransactionEntry, extraction_data: Optional[dict] = None):
        """Append a transaction to the Excel file."""
        try:
            wb = openpyxl.load_workbook(self.output_path)
            ws = wb.active
            
            # Get next row
            next_row = ws.max_row + 1
            
            # Prepare row data
            row_data = [
                transaction.date,
                transaction.transaction_id,
                extraction_data.get("invoice_id") if extraction_data else None,
                extraction_data.get("vendor_name") if extraction_data else None,
                transaction.description,
                transaction.debit_account,
                transaction.debit_amount,
                transaction.credit_account,
                transaction.credit_amount,
                transaction.tds_account,
                transaction.tds_amount,
                transaction.gst_amount,
                f"{transaction.confidence_score:.2%}",
                transaction.rule_applied or "",
                transaction.status,
            ]
            
            # Write row
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=next_row, column=col_idx, value=value)
                
                # Apply formatting
                self._format_cell(cell, col_idx)
            
            # Style status based on value
            status_cell = ws.cell(row=next_row, column=15)  # Status column
            if transaction.status == "flagged" or transaction.status == "needs_review":
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif transaction.status == "approved":
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            wb.save(self.output_path)
            logger.info(f"Transaction {transaction.transaction_id} appended to Excel")
            
        except Exception as e:
            logger.error(f"Error appending transaction to Excel: {str(e)}")
            raise
    
    def _format_cell(self, cell, col_idx: int):
        """Apply formatting to a cell based on column."""
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = thin_border
        
        # Format numbers
        if col_idx in [7, 9, 11, 12]:  # Amount columns
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
        
        # Format percentage
        if col_idx == 13:  # Confidence column
            cell.alignment = Alignment(horizontal="center")
        
        # Center align
        if col_idx in [1, 2, 13, 15]:  # Date, Transaction ID, Confidence, Status
            cell.alignment = Alignment(horizontal="center")
    
    def update_transaction(self, transaction_id: str, updates: dict):
        """Update an existing transaction row."""
        try:
            wb = openpyxl.load_workbook(self.output_path)
            ws = wb.active
            
            # Find transaction row
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[1].value == transaction_id:  # Column B is Transaction ID
                    # Update values
                    if "status" in updates:
                        row[14].value = updates["status"]  # Column O
                    if "rule_applied" in updates:
                        row[13].value = updates["rule_applied"]  # Column N
                    if "confidence" in updates:
                        row[12].value = f"{updates['confidence']:.2%}"  # Column M
                    
                    wb.save(self.output_path)
                    logger.info(f"Transaction {transaction_id} updated in Excel")
                    return True
            
            logger.warning(f"Transaction {transaction_id} not found in Excel")
            return False
        
        except Exception as e:
            logger.error(f"Error updating transaction in Excel: {str(e)}")
            raise
    
    def get_summary(self) -> dict:
        """Get summary statistics from ledger."""
        try:
            df = pd.read_excel(self.output_path)
            
            summary = {
                "total_transactions": len(df),
                "total_debit": df["Debit Amount"].sum(),
                "total_credit": df["Credit Amount"].sum(),
                "total_tds": df["TDS Amount"].sum(),
                "avg_confidence": df["Confidence"].str.rstrip('%').astype(float).mean(),
                "status_breakdown": df["Status"].value_counts().to_dict(),
            }
            
            return summary
        except Exception as e:
            logger.error(f"Error generating summary: {str(e)}")
            return {}
